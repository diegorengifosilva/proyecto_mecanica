# -*- coding: utf-8 -*-
"""
Módulo: reports.py (Dentro de simulador_api)
Descripción: Genera los reportes estructurados para descarga desde la web para los 5 escenarios.
"""
import csv
import io
from datetime import datetime

def generar_reporte_txt(info, pasos):
    """
    Genera un string formateado con el reporte académico de la simulación.
    """
    output = io.StringIO()
    
    output.write("=========================================================================\n")
    output.write("       REPORTE ACADÉMICO DE SIMULACIÓN - SEGUNDA LEY DE NEWTON          \n")
    output.write("                 UNIVERSIDAD TECNOLÓGICA DEL PERÚ                        \n")
    output.write("=========================================================================\n\n")
    
    fecha_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    output.write(f"Fecha de Reporte: {fecha_str}\n")
    output.write(f"ID del Ensayo: {info.get('id', 'Temporal')}\n")
    
    escenario_nombre = info.get('escenario', 'Automovil')
    output.write(f"Escenario Simulado: {escenario_nombre}\n")
    if info.get('circuito'):
        output.write(f"Circuito / Trazado 3D: {info.get('circuito')}\n")
    output.write("\n")
    
    output.write("-------------------------------------------------------------------------\n")
    output.write("1. PARÁMETROS DE ENTRADA (CONDICIONES INICIALES)\n")
    output.write("-------------------------------------------------------------------------\n")
    output.write(f"  - Masa (m): {info.get('masa', 0.0):.2f} kg\n")
    
    escL = escenario_nombre.lower()
    if escL in ["automovil", "automotriz", "camion", "motocicleta"]:
        output.write(f"  - Velocidad Inicial (v0): {info.get('v_inicial', 0.0):.2f} m/s\n")
        output.write(f"  - Fuerza Aplicada (F): {info.get('fuerza', 0.0):.2f} N\n")
        output.write(f"  - Coeficiente de Fricción (mu): {info.get('friccion', 0.0):.4f}\n")
        if info.get('terreno'):
            output.write(f"  - Clima / Carretera: {info.get('terreno')}\n")
        else:
            output.write(f"  - Clima / Carretera: Asfalto {info.get('clima', 'Seco')}\n")
        output.write(f"  - Gravedad (g): 9.81 m/s²\n")
        
    elif escL == "elevador":
        output.write(f"  - Tensión del Cable (T): {info.get('fuerza', 0.0):.2f} N\n")
        output.write(f"  - Coeficiente de Arrastre (b): {info.get('friccion', 0.0):.4f} N*s/m\n")
        if info.get('entorno'):
            output.write(f"  - Fluido del Medio: {info.get('entorno')}\n")
        else:
            output.write(f"  - Fluido del Medio: {info.get('clima', 'Aire')}\n")
        output.write(f"  - Gravedad (g): 9.81 m/s²\n")
        
    elif escL == "avion":
        output.write(f"  - Empuje Turbina (F_empuje): {info.get('fuerza', 0.0):.2f} N\n")
        output.write(f"  - Coeficiente de Fricción de Pista (mu): {info.get('friccion', 0.0):.4f}\n")
        output.write(f"  - Gravedad (g): 9.81 m/s²\n")
        output.write(f"  - Pista Clima: {info.get('clima', 'Pista Seca')}\n")
        
    elif escL == "atwood":
        output.write(f"  - Masa Izquierda (m1): {info.get('masa', 0.0):.2f} kg\n")
        output.write(f"  - Masa Derecha (m2): {info.get('masa_2', 0.0):.2f} kg\n")
        output.write(f"  - Gravedad (g): 9.81 m/s²\n")
        output.write(f"  - Perfil de Fuerza: {info.get('perfil_fuerza', 'Constante')}\n")
    elif escL == "curva":
        output.write(f"  - Radio de Curva (R): {info.get('altura', info.get('radio_curva', 80.0)):.2f} m\n")
        output.write(f"  - Ángulo de Peralte (θ): {info.get('inclinacion', 0.0):.2f}°\n")
        output.write(f"  - Coeficiente de Fricción Estática (mu_s): {info.get('friccion_estatica', 0.80):.2f}\n")
        output.write(f"  - Coeficiente de Fricción Dinámica (mu_k): {info.get('friccion', 0.00):.4f}\n")
        output.write(f"  - Fuerza de Tracción / Freno (F): {info.get('fuerza', 0.0):.2f} N\n")
        output.write(f"  - Velocidad Inicial (v0): {info.get('v_inicial', 0.0):.2f} m/s\n")
        output.write(f"  - Gravedad (g): 9.81 m/s²\n")
        output.write(f"  - Perfil de Fuerza: {info.get('perfil_fuerza', 'Constante')}\n")

    output.write("\n")
    
    output.write("-------------------------------------------------------------------------\n")
    output.write("2. RESULTADOS OBTENIDOS (MAGNITUDES FÍSICAS RESULTANTES)\n")
    output.write("-------------------------------------------------------------------------\n")
    output.write(f"  - Tiempo Total de Simulación: {info.get('tiempo_total', 0.0):.3f} s\n")
    output.write(f"  - Desplazamiento/Altura Máxima: {info.get('distancia_recorrida', 0.0):.3f} m\n")
    output.write(f"  - Aceleración Promedio Calculada: {info.get('aceleracion_promedio', 0.0):.4f} m/s²\n")
    output.write("\n")
    
    output.write("-------------------------------------------------------------------------\n")
    output.write("3. MARCO TEÓRICO Y FÓRMULAS DE APLICACIÓN\n")
    output.write("-------------------------------------------------------------------------\n")
    
    if escL in ["automovil", "automotriz", "camion", "motocicleta"]:
        output.write("  Ecuaciones del Movimiento Rectilíneo con Fricción de Carretera y Arrastre de Aire:\n")
        output.write("    * Fuerza Normal: N = m * g\n")
        output.write("    * Fuerza de Fricción Dinámica: fk = mu * N\n")
        output.write("    * Resistencia Aerodinámica: Fd = 0.5 * rho * Cd * A * v²\n")
        output.write("    * Fuerza Neta Horizontal:\n")
        output.write("        - Aceleración: ΣFx = F_motor - fk - Fd\n")
        output.write("        - Frenado: ΣFx = -(F_frenado + fk + Fd)\n")
        output.write("    * Segunda Ley de Newton: a = ΣFx / m\n")
        
    elif escL == "elevador":
        output.write("  Ecuaciones para Movimiento Vertical con Gravedad y Resistencia Viscosa:\n")
        output.write("    * Fuerza de Gravedad (Peso): P = m * g\n")
        output.write("    * Fuerza de Arrastre (Fluido): Fd = - b * v\n")
        output.write("    * Fuerza Neta Vertical: ΣFy = Tensión - Peso + Fd\n")
        output.write("    * Segunda Ley de Newton (Aceleración): a = ΣFy / m\n")
        
    elif escL == "avion":
        output.write("  Ecuaciones de Despegue en Pista con Sustentación Aerodinámica:\n")
        output.write("    * Fuerza de Gravedad (Peso): P = m * g\n")
        output.write("    * Fuerza de Sustentación Vertical: L = Cl * v²\n")
        output.write("    * Fuerza Normal (sobre la pista): N = max(0, P - L)\n")
        output.write("    * Fricción de Neumáticos: fr = mu * N\n")
        output.write("    * Arrastre Aerodinámico: Fd = b * v²\n")
        output.write("    * Fuerza Neta Horizontal: ΣFx = Empuje - fr - Fd\n")
        output.write("    * Segunda Ley de Newton (Aceleración): a = ΣFx / m\n")
        output.write("    * NOTA: Cuando L >= P, el avión despega, la Normal se hace 0 N y fr = 0 N.\n")
        
    elif escL == "atwood":
        output.write("  Ecuaciones de la Máquina de Atwood (Doble Masa y Polea Fija):\n")
        output.write("    * Fuerzas de Gravedad: P1 = m1 * g , P2 = m2 * g\n")
        output.write("    * Fuerza Neta del Sistema: ΣF = |m2 - m1| * g\n")
        output.write("    * Inercia Total del Sistema: M_total = m1 + m2\n")
        output.write("    * Segunda Ley de Newton (Aceleración): a = ΣF / M_total\n")
        output.write("    * Tensión en la Cuerda (T): T = 2 * m1 * m2 * g / (m1 + m2)\n")
    elif escL == "curva":
        output.write("  Ecuaciones de Movimiento en Curva Peraltada (Leyes de Newton y Fuerza Centrípeta):\n")
        output.write("    * Fuerza Normal Dinámica: N = m * (g * cos(θ) + v² / R * sen(θ))\n")
        output.write("    * Fricción Lateral Requerida: fs_req = m * (v² / R * cos(θ) - g * sen(θ))\n")
        output.write("    * Límites de Adherencia (Sin Derrape/Deslizamiento):\n")
        output.write("        - v_max = √[ g * R * (tan(θ) + μs) / (1 - μs * tan(θ)) ]\n")
        output.write("        - v_min = √[ g * R * (tan(θ) - μs) / (1 + μs * tan(θ)) ] (si tan(θ) > μs)\n")
        output.write("    * Fricción Cinética Tangencial: fk = μk * N\n")
        output.write("    * Aceleración Tangencial: a = (F_motor - fk) / m o a = -(F_freno + fk) / m\n")

    output.write("\n")
    
    output.write("-------------------------------------------------------------------------\n")
    output.write("4. TABLA DE MUESTREO TEMPORAL (SEGUNDO A SEGUNDO)\n")
    output.write("-------------------------------------------------------------------------\n")
    output.write(f"{'Tiempo (s)':<12}{'Posición (m)':<15}{'Velocidad (m/s)':<18}{'Aceleración (m/s²)':<20}{'Fuerza Neta (N)':<15}\n")
    output.write("-" * 80 + "\n")
    
    total_pasos = len(pasos)
    pasos_imprimir = []
    if total_pasos > 0:
        pasos_imprimir.append(pasos[0])
        # Imprimir una muestra representativa de hasta 30 pasos para no saturar el TXT
        intervalo = max(1, total_pasos // 30)
        for i in range(1, total_pasos - 1):
            if i % intervalo == 0:
                pasos_imprimir.append(pasos[i])
        pasos_imprimir.append(pasos[-1])
        
    vistas = set()
    for paso in pasos_imprimir:
        t_str = f"{paso['t']:.2f}"
        if t_str not in vistas:
            vistas.add(t_str)
            f_neta = paso.get('fuerza_neta', paso.get('fuerzas', {}).get('FuerzaNeta', 0.0))
            output.write(
                f"{paso['t']:<12.2f}"
                f"{paso['x']:<15.4f}"
                f"{paso['v']:<18.4f}"
                f"{paso['a']:<20.4f}"
                f"{f_neta:<15.1f}\n"
            )
            
    output.write("\n=========================================================================\n")
    output.write("       FIN DEL REPORTE ACADÉMICO - CÁTEDRA DE FÍSICA UTP                 \n")
    output.write("=========================================================================\n")
    
    return output.getvalue()

def generar_reporte_csv(pasos):
    """
    Genera un string en formato CSV con el paso a paso detallado.
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(["Tiempo (s)", "Posicion (m)", "Velocidad (m/s)", "Aceleracion (m/s2)", "Fuerza Neta (N)"])
    
    for paso in pasos:
        f_neta = paso.get('fuerza_neta', paso.get('fuerzas', {}).get('FuerzaNeta', 0.0))
        writer.writerow([
            f"{paso['t']:.3f}",
            f"{paso['x']:.4f}",
            f"{paso['v']:.4f}",
            f"{paso['a']:.4f}",
            f"{f_neta:.2f}"
        ])
        
    return output.getvalue()


def generar_reporte_excel(info, pasos):
    """
    Genera un archivo Excel (.xlsx) estructurado en dos pestañas,
    con diseño premium en colores azul marino y celeste, incluyendo
    fórmulas activas de Excel para los cálculos de error y diferencias.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ----------------------------------------------------
    # PESTAÑA 1: Resumen y Fórmulas
    # ----------------------------------------------------
    ws1 = wb.create_sheet(title="Resumen del Ensayo")
    ws1.views.sheetView[0].showGridLines = True

    # Definir paleta de colores y estilos
    navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")  # Azul Oscuro Slate
    light_blue_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Azul Marino Real
    accent_blue_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Azul Celeste
    ice_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")  # Celeste muy claro para etiquetas
    
    white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
    normal_font = Font(name="Segoe UI", size=11, color="334155")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Título Principal
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "LABORATORIO NEWTONIANO INTERACTIVO"
    ws1["A1"].font = title_font
    ws1["A1"].fill = navy_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:G2")
    ws1["A2"] = "Reporte Científico de Ensayo - Segunda Ley de Newton (UTP)"
    ws1["A2"].font = subtitle_font
    ws1["A2"].fill = light_blue_fill
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 25

    # Bloque: Datos Generales
    ws1.merge_cells("A4:C4")
    ws1["A4"] = "DATOS GENERALES DEL ENSAYO"
    ws1["A4"].font = white_font
    ws1["A4"].fill = accent_blue_fill
    ws1["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws1["A4"].border = thin_border
    
    datos_generales = [
        ("ID del Ensayo", info.get('id', 'Temporal')),
        ("Fecha / Hora", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ("Escenario Simulado", info.get('escenario', 'Automóvil')),
    ]
    if info.get('circuito'):
        datos_generales.append(("Circuito / Trazado 3D", info.get('circuito')))
    datos_generales.extend([
        ("Clima / Entorno", info.get('clima', info.get('terreno', info.get('entorno', 'Seco')))),
        ("Perfil de Fuerza", info.get('perfil_fuerza', 'Constante'))
    ])
    
    for idx, (label, val) in enumerate(datos_generales, start=5):
        ws1.cell(row=idx, column=1, value=label).font = bold_font
        ws1.cell(row=idx, column=1).fill = ice_blue_fill
        ws1.cell(row=idx, column=1).border = thin_border
        
        ws1.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=3)
        val_cell = ws1.cell(row=idx, column=2, value=val)
        val_cell.font = normal_font
        val_cell.border = thin_border
        ws1.cell(row=idx, column=3).border = thin_border

    # Calcular métricas de error promedio
    valid_pasos = [p for p in pasos if p.get('t', 0.0) > 0]
    if valid_pasos:
        acel_real_prom = sum(p.get('a', 0.0) for p in valid_pasos) / len(valid_pasos)
        acel_teorica_prom = sum(p.get('a_teorica', 0.0) for p in valid_pasos) / len(valid_pasos)
        
        ultimo_paso = pasos[-1]
        vel_real_final = ultimo_paso.get('v', 0.0)
        vel_teorica_final = ultimo_paso.get('v_teorica', 0.0)
        
        pos_real_final = ultimo_paso.get('x', 0.0)
        pos_teorica_final = ultimo_paso.get('x_teorica', 0.0)
        
        error_acel = (abs(acel_teorica_prom - acel_real_prom) / abs(acel_teorica_prom) * 100) if abs(acel_teorica_prom) > 1e-4 else 0.0
    else:
        acel_real_prom = info.get('aceleracion_promedio', 0.0)
        acel_teorica_prom = info.get('aceleracion_promedio', 0.0)
        vel_real_final = 0.0
        vel_teorica_final = 0.0
        pos_real_final = info.get('distancia_recorrida', 0.0)
        pos_teorica_final = info.get('distancia_recorrida', 0.0)
        error_acel = 0.0

    # Bloque: Resultados
    ws1.merge_cells("E4:G4")
    ws1["E4"] = "RESULTADOS CLAVE DE LA SIMULACIÓN"
    ws1["E4"].font = white_font
    ws1["E4"].fill = accent_blue_fill
    ws1["E4"].alignment = Alignment(horizontal="left", vertical="center")
    ws1["E4"].border = thin_border
    
    resultados = [
        ("Tiempo Total de Simulación", info.get('tiempo_total', 0.0), "s"),
        ("Desplazamiento / Altura Máx", info.get('distancia_recorrida', 0.0), "m"),
        ("Aceleración Promedio Real", acel_real_prom, "m/s²"),
        ("Aceleración Prom. Teórica", acel_teorica_prom, "m/s²"),
        ("Error Relativo Aceleración", error_acel / 100.0, "%")
    ]
    
    for idx, (label, val, unit) in enumerate(resultados, start=5):
        ws1.cell(row=idx, column=5, value=label).font = bold_font
        ws1.cell(row=idx, column=5).fill = ice_blue_fill
        ws1.cell(row=idx, column=5).border = thin_border
        
        val_cell = ws1.cell(row=idx, column=6, value=val)
        val_cell.font = normal_font
        val_cell.border = thin_border
        
        if unit == "%":
            val_cell.number_format = '0.0%'
        elif unit == "s":
            val_cell.number_format = '0.000'
        elif unit == "m":
            val_cell.number_format = '0.00'
        else:
            val_cell.number_format = '0.0000'
            
        unit_cell = ws1.cell(row=idx, column=7, value=unit)
        unit_cell.font = normal_font
        unit_cell.border = thin_border

    # Bloque: Parámetros de Entrada
    ws1.merge_cells("A11:C11")
    ws1["A11"] = "PARÁMETROS FÍSICOS CONFIGURADOS"
    ws1["A11"].font = white_font
    ws1["A11"].fill = accent_blue_fill
    ws1["A11"].alignment = Alignment(horizontal="left", vertical="center")
    ws1["A11"].border = thin_border
    
    escL = info.get('escenario', 'Automovil').lower()
    
    parametros = [
        ("Masa Principal (m1 / Cabina)", info.get('masa', 0.0), "kg")
    ]
    if escL == 'atwood':
        parametros.append(("Masa Secundaria (m2 / Contrapeso)", info.get('masa_2', 0.0), "kg"))
    
    if escL == 'elevador':
        parametros.append(("Tensión Nominal del Cable (T)", info.get('fuerza', 0.0), "N"))
        parametros.append(("Coeficiente de Resistencia Viscosa (b)", info.get('friccion', 0.0), "N·s/m"))
    elif escL == 'avion':
        parametros.append(("Empuje de los Turborreactores (F)", info.get('fuerza', 0.0), "N"))
        parametros.append(("Coef. Fricción en Pista (mu)", info.get('friccion', 0.0), ""))
        parametros.append(("Coef. Resistencia del Aire (b)", info.get('resistencia_aire', 0.0), ""))
    elif escL == 'curva':
        parametros.append(("Radio de Curva (R)", info.get('altura', info.get('radio_curva', 80.0)), "m"))
        parametros.append(("Ángulo de Peralte (θ)", info.get('inclinacion', 0.0), "°"))
        parametros.append(("Coef. Fricción Estática (mu_s)", info.get('friccion_estatica', 0.8), ""))
        parametros.append(("Coef. Fricción Cinética (mu_k)", info.get('friccion', 0.0), ""))
        parametros.append(("Fuerza de Tracción / Freno (F)", info.get('fuerza', 0.0), "N"))
        parametros.append(("Velocidad Inicial (v0)", info.get('v_inicial', 0.0), "m/s"))
    else:
        # Terrestre
        parametros.append(("Fuerza de Tracción / Freno (F)", info.get('fuerza', 0.0), "N"))
        parametros.append(("Coef. Fricción Cinética (mu_k)", info.get('friccion', 0.0), ""))
        parametros.append(("Coef. Fricción Estática (mu_s)", info.get('friccion_estatica', 0.0), ""))
        parametros.append(("Coef. Resistencia de Aire (Cd*A)", info.get('resistencia_aire', 0.0), ""))
        parametros.append(("Pendiente de Rampa (inclinación)", info.get('inclinacion', 0.0), "°"))
        parametros.append(("Velocidad Inicial (v0)", info.get('v_inicial', 0.0), "m/s"))
        
    for idx, (label, val, unit) in enumerate(parametros, start=12):
        ws1.cell(row=idx, column=1, value=label).font = bold_font
        ws1.cell(row=idx, column=1).border = thin_border
        
        val_cell = ws1.cell(row=idx, column=2, value=val)
        val_cell.font = normal_font
        val_cell.border = thin_border
        if isinstance(val, (int, float)):
            if val == int(val):
                val_cell.number_format = '#,##0'
            else:
                val_cell.number_format = '0.0000'
                
        unit_cell = ws1.cell(row=idx, column=3, value=unit)
        unit_cell.font = normal_font
        unit_cell.border = thin_border

    # Bloque: Marco Teórico y Ecuaciones
    ws1.merge_cells("E11:G11")
    ws1["E11"] = "MARCO TEÓRICO Y ECUACIONES"
    ws1["E11"].font = white_font
    ws1["E11"].fill = accent_blue_fill
    ws1["E11"].alignment = Alignment(horizontal="left", vertical="center")
    ws1["E11"].border = thin_border
    
    ecuaciones = []
    if escL in ["automovil", "automotriz", "camion", "motocicleta"]:
        ecuaciones = [
            "Segunda Ley de Newton con Fuerzas de Fricción y Arrastre",
            "Fuerza Normal: N = m * g * cos(θ)",
            "Fricción Cinética: fk = μk * N",
            "Fricción Estática Máx: fs_max = μs * N",
            "Resistencia del Aire: Fd = 0.5 * ρ * Cd * A * v²",
            "Ecuación del Movimiento Real:",
            "  a = [F_motor - fk - Fd - m * g * sin(θ)] / m",
            "Modelo Teórico Ideal (Sin Arrastre, b = 0):",
            "  a_teorica = [F_motor - fk - m * g * sin(θ)] / m"
        ]
    elif escL == 'elevador':
        ecuaciones = [
            "Movimiento Vertical con Resistencia Viscosa",
            "Fuerza Peso: P = m * g",
            "Resistencia del Fluido (Arrastre): Fd = - b * v",
            "Ecuación del Movimiento Real:",
            "  a = [Tensión - Peso - b * v] / m",
            "Modelo Teórico Ideal (En Vacío, b = 0):",
            "  a_teorica = [Tensión - Peso] / m"
        ]
    elif escL == 'avion':
        ecuaciones = [
            "Despegue con Sustentación Aerodinámica Dinámica",
            "Fuerza Peso: P = m * g * cos(θ)",
            "Sustentación Aerodinámica: Sustentación = Cl * v²",
            "Fuerza Normal en Pista: N = max(0, Peso - Sustentación)",
            "Arrastre del Aire: Fd = b * v²",
            "Ecuación del Movimiento Real:",
            "  a = [Empuje - μk * N - Fd - m * g * sin(θ)] / m",
            "Modelo Teórico Ideal (Sin Arrastre, b = 0):",
            "  a_teorica = [Empuje - μk * N - m * g * sin(θ)] / m"
        ]
    elif escL == 'atwood':
        ecuaciones = [
            "Dinámica de Sistema Acoplado (Máquina de Atwood)",
            "Pesos Individuales: P1 = m1 * g , P2 = m2 * g",
            "Fuerza Neta Aceleradora: ΔF = |m2 - m1| * g",
            "Inercia Total del Sistema: M_total = m1 + m2",
            "Ecuación del Movimiento Real:",
            "  a = [ΔF - b * v] / M_total",
            "Modelo Teórico Ideal (Sin Arrastre, b = 0):",
            "  a_teorica = ΔF / M_total",
            "Tensión Ideal del Cable:",
            "  T = 2 * m1 * m2 * g / M_total"
        ]
    elif escL == 'curva':
        ecuaciones = [
            "Dinámica de Vehículo en Curva Peraltada",
            "Fuerza Normal Dinámica: N = m * [g * cos(θ) + v² / R * sen(θ)]",
            "Fricción Lateral Requerida: fs_req = m * [v² / R * cos(θ) - g * sen(θ)]",
            "Fuerza de Fricción Dinámica (Longitudinal): fk = μk * N",
            "Límites de Adherencia (Deslizarse / Derrapar):",
            "  v_max = sqrt[ g * R * (tan(θ) + μs) / (1 - μs * tan(θ)) ]",
            "  v_min = sqrt[ g * R * (tan(θ) - μs) / (1 + μs * tan(θ)) ] (si tan(θ) > μs)",
            "Ecuación del Movimiento Tangencial Real:",
            "  a = [F_motor - fk - Fd] / m"
        ]
        
    for idx, eq in enumerate(ecuaciones, start=12):
        ws1.merge_cells(start_row=idx, start_column=5, end_row=idx, end_column=7)
        eq_cell = ws1.cell(row=idx, column=5, value=eq)
        eq_cell.font = normal_font
        eq_cell.border = thin_border
        ws1.cell(row=idx, column=6).border = thin_border
        ws1.cell(row=idx, column=7).border = thin_border

    # Bloque: Comparación Final
    row_start_comp = max(12 + len(parametros), 12 + len(ecuaciones)) + 2
    ws1.merge_cells(start_row=row_start_comp, start_column=1, end_row=row_start_comp, end_column=7)
    comp_header = ws1.cell(row=row_start_comp, column=1, value="COMPARATIVA DE MODELOS - ESTADO FINAL")
    comp_header.font = white_font
    comp_header.fill = light_blue_fill
    comp_header.alignment = Alignment(horizontal="left", vertical="center")
    comp_header.border = thin_border
    
    headers_comp = ["Magnitud Física", "Valor Real (Simulado)", "Valor Teórico (Ideal)", "Desviación Absoluta", "Error Relativo"]
    row_headers = row_start_comp + 1
    
    ws1.cell(row=row_headers, column=1, value=headers_comp[0]).font = bold_font
    ws1.cell(row=row_headers, column=1).border = thin_border
    ws1.cell(row=row_headers, column=1).fill = ice_blue_fill
    
    ws1.merge_cells(start_row=row_headers, start_column=2, end_row=row_headers, end_column=3)
    ws1.cell(row=row_headers, column=2, value=headers_comp[1]).font = bold_font
    ws1.cell(row=row_headers, column=2).border = thin_border
    ws1.cell(row=row_headers, column=2).fill = ice_blue_fill
    ws1.cell(row=row_headers, column=3).border = thin_border
    ws1.cell(row=row_headers, column=3).fill = ice_blue_fill
    
    ws1.cell(row=row_headers, column=4, value=headers_comp[2]).font = bold_font
    ws1.cell(row=row_headers, column=4).border = thin_border
    ws1.cell(row=row_headers, column=4).fill = ice_blue_fill
    
    ws1.cell(row=row_headers, column=5, value=headers_comp[3]).font = bold_font
    ws1.cell(row=row_headers, column=5).border = thin_border
    ws1.cell(row=row_headers, column=5).fill = ice_blue_fill
    
    ws1.merge_cells(start_row=row_headers, start_column=6, end_row=row_headers, end_column=7)
    ws1.cell(row=row_headers, column=6, value=headers_comp[4]).font = bold_font
    ws1.cell(row=row_headers, column=6).border = thin_border
    ws1.cell(row=row_headers, column=6).fill = ice_blue_fill
    ws1.cell(row=row_headers, column=7).border = thin_border
    ws1.cell(row=row_headers, column=7).fill = ice_blue_fill
    
    comp_data = [
        ("Aceleración Promedio (m/s²)", acel_real_prom, acel_teorica_prom, "acel"),
        ("Velocidad Final (m/s)", vel_real_final, vel_teorica_final, "vel"),
        ("Desplazamiento Final (m)", pos_real_final, pos_teorica_final, "pos")
    ]
    
    for offset, (label, real_val, teor_val, tag) in enumerate(comp_data):
        curr_row = row_headers + 1 + offset
        
        # Etiqueta de la Magnitud
        ws1.cell(row=curr_row, column=1, value=label).font = bold_font
        ws1.cell(row=curr_row, column=1).border = thin_border
        
        # Valor Real (Simulado)
        ws1.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
        r_cell = ws1.cell(row=curr_row, column=2, value=real_val)
        r_cell.font = normal_font
        r_cell.border = thin_border
        r_cell.number_format = '0.0000'
        ws1.cell(row=curr_row, column=3).border = thin_border
        
        # Valor Teórico (Ideal)
        t_cell = ws1.cell(row=curr_row, column=4, value=teor_val)
        t_cell.font = normal_font
        t_cell.border = thin_border
        t_cell.number_format = '0.0000'
        
        # Desviación Absoluta (Fórmula)
        d_cell = ws1.cell(row=curr_row, column=5, value=f"=ABS(B{curr_row}-D{curr_row})")
        d_cell.font = normal_font
        d_cell.border = thin_border
        d_cell.number_format = '0.0000'
        
        # Error Relativo (Fórmula)
        ws1.merge_cells(start_row=curr_row, start_column=6, end_row=curr_row, end_column=7)
        e_cell = ws1.cell(row=curr_row, column=6, value=f"=IF(ABS(D{curr_row})>1E-4, ABS(B{curr_row}-D{curr_row})/ABS(D{curr_row}), 0.0)")
        e_cell.font = normal_font
        e_cell.border = thin_border
        e_cell.number_format = '0.0%'
        ws1.cell(row=curr_row, column=7).border = thin_border

    # Bloque: Caja Explicativa de Error
    row_expl = row_headers + 5
    ws1.merge_cells(start_row=row_expl, start_column=1, end_row=row_expl+2, end_column=7)
    
    b_val = float(info.get('resistencia_aire', 0.0))
    if escL == 'atwood':
        if b_val > 0:
            expl_text = f"Análisis de Desviación Académica: El margen de error del sistema acoplado se debe a la resistencia del aire (b = {b_val:.2f} N·s/m). El modelo ideal no tiene pérdidas disipativas, asumiendo cuerdas inextensibles sin fricción ni masa."
        else:
            expl_text = "Análisis de Desviación Académica: Con resistencia de aire nula (b = 0), el comportamiento físico simulado es perfectamente coincidente con la ecuación teórica de la Máquina de Atwood."
    elif escL == 'elevador':
        if b_val > 0:
            expl_text = f"Análisis de Desviación Académica: La diferencia en la velocidad surge por la fricción viscosa del medio (b = {b_val:.2f} N·s/m) que ejerce una fuerza disipativa contraria al movimiento proporcional a la velocidad, a diferencia del modelo teórico en el vacío."
        else:
            expl_text = "Análisis de Desviación Académica: Al no existir resistencia de fluido (vacío), la cabina sigue de manera ideal la aceleración neta constante a = (T - P)/m."
    elif escL == 'curva':
        if b_val > 0:
            expl_text = f"Análisis de Desviación Académica: El margen de error surge por el arrastre aerodinámico (b = {b_val:.2f}) que actúa en dirección opuesta al movimiento tangencial de avance en la curva peraltada."
        else:
            expl_text = "Análisis de Desviación Académica: Con resistencia de aire nula (b = 0), el comportamiento real coincide plenamente con las ecuaciones teóricas de Newton para una curva peraltada estable."
    else:
        if b_val > 0:
            expl_text = f"Análisis de Desviación Académica: La resistencia aerodinámica (Fd = 0.5·ρ·Cd·A·v²) que crece de forma cuadrática con la velocidad es la causa del error relativo. Genera una fuerza de retardo progresiva que reduce la aceleración real frente a la aceleración teórica constante."
        else:
            expl_text = "Análisis de Desviación Académica: Al configurarse resistencia de aire nula (Cd*A = 0), el movimiento terrestre se rige rigurosamente por la fuerza motriz, la componente del peso en rampa y la fricción constante del suelo."
            
    expl_cell = ws1.cell(row=row_expl, column=1, value=expl_text)
    expl_cell.font = Font(name="Segoe UI", size=10, italic=True, color="475569")
    expl_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Aplicar bordes al cuadro explicativo
    for r in range(row_expl, row_expl+3):
        for c in range(1, 8):
            cell = ws1.cell(row=r, column=c)
            cell.border = thin_border
            if r == row_expl and c == 1:
                pass
            else:
                cell.value = None

    # Ajustar anchos de columna de la Pestaña 1
    ws1.column_dimensions['A'].width = 34
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 6
    ws1.column_dimensions['E'].width = 28
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 12


    # ----------------------------------------------------
    # PESTAÑA 2: Muestreo Detallado Paso a Paso
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Muestreo Temporal")
    ws2.views.sheetView[0].showGridLines = True

    # Cabecera Pestaña 2
    ws2.merge_cells("A1:J1")
    ws2["A1"] = "REGISTRO TEMPORAL DE SIMULACIÓN - PASO A PASO"
    ws2["A1"].font = title_font
    ws2["A1"].fill = navy_fill
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    headers_ws2 = [
        "Tiempo (s)", 
        "Posición Real (m)", 
        "Velocidad Real (m/s)", 
        "Aceleración Real (m/s²)", 
        "Fuerza Neta Real (N)",
        "Posición Teórica (m)",
        "Velocidad Teórica (m/s)",
        "Aceleración Teórica (m/s²)",
        "Diferencia Velocidad (m/s)",
        "Diferencia Aceleración (m/s²)"
    ]
    
    ws2.row_dimensions[2].height = 28
    for col_idx, h in enumerate(headers_ws2, start=1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font = white_font
        cell.fill = light_blue_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Escribir filas paso a paso
    for idx, paso in enumerate(pasos, start=3):
        t = paso.get('t', 0.0)
        x = paso.get('x', 0.0)
        v = paso.get('v', 0.0)
        a = paso.get('a', 0.0)
        f_neta = paso.get('fuerza_neta', paso.get('fuerzas', {}).get('Fuerza Neta', paso.get('fuerzas', {}).get('FuerzaNeta', 0.0)))
        
        x_teor = paso.get('x_teorica', 0.0)
        v_teor = paso.get('v_teorica', 0.0)
        a_teor = paso.get('a_teorica', 0.0)
        
        row_vals = [
            t, x, v, a, f_neta, 
            x_teor, v_teor, a_teor,
            f"=ABS(C{idx}-G{idx})",
            f"=ABS(D{idx}-H{idx})"
        ]
        
        ws2.row_dimensions[idx].height = 20
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            
            # Formatos numéricos y alineaciones
            if col_idx == 1:
                cell.number_format = '0.000'
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [2, 3, 6, 7, 9]:
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [4, 8, 10]:
                cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Ajustar automáticamente anchos de columna de la Pestaña 2
    for col in ws2.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.coordinate in ws2.merged_cells:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # Escribir en buffer de bytes in-memory
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel_historial(records):
    """
    Genera un archivo Excel (.xlsx) con el historial completo de ensayos persistidos,
    con diseño premium en colores azul marino y celeste, según la estructura solicitada.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Ensayos"
    ws.views.sheetView[0].showGridLines = True

    # Paleta de colores e incidentes
    dark_blue_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")  # Azul Acero Oscuro
    medium_blue_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")  # Azul del encabezado
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Fondo azul suave para KPIs
    zebra_fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")  # Gris azulado tenue para filas alternas

    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="595959")
    bold_dark_font = Font(name="Segoe UI", size=11, bold=True, color="1B365D")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_bold_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
    data_normal_font = Font(name="Segoe UI", size=11, color="000000")
    kpi_label_font = Font(name="Segoe UI", size=11, bold=True, color="333333")
    kpi_value_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # 1. Cabecera del Título (Filas 1-2)
    ws.merge_cells("A1:I1")
    ws["A1"] = "ANÁLISIS DE LA SEGUNDA LEY DE NEWTON - REGISTRO HISTÓRICO"
    ws["A1"].font = title_font
    ws["A1"].fill = dark_blue_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:I2")
    ws["A2"] = "Proyecto de Mecánica Clásica - Persistencia de Ensayos de Simulación"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # 2. Bloque de KPI (Filas 4-5)
    total_records = len(records)
    max_masa = max([r.masa for r in records]) if total_records > 0 else 0.0
    max_fuerza = max([r.fuerza for r in records]) if total_records > 0 else 0.0

    ws.merge_cells("B4:C4")
    ws.cell(row=4, column=2, value="Total Ensayos").font = kpi_label_font
    ws.cell(row=4, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=2).fill = zebra_fill
    ws.cell(row=4, column=2).border = thin_border
    ws.cell(row=4, column=3).border = thin_border
    
    ws.merge_cells("B5:C5")
    val_total = ws.cell(row=5, column=2, value=total_records)
    val_total.font = kpi_value_font
    val_total.alignment = Alignment(horizontal="center", vertical="center")
    val_total.fill = light_blue_fill
    val_total.border = thin_border
    ws.cell(row=5, column=3).border = thin_border

    ws.merge_cells("E4:F4")
    ws.cell(row=4, column=5, value="Masa Máx (kg)").font = kpi_label_font
    ws.cell(row=4, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=5).fill = zebra_fill
    ws.cell(row=4, column=5).border = thin_border
    ws.cell(row=4, column=6).border = thin_border
    
    ws.merge_cells("E5:F5")
    val_masa = ws.cell(row=5, column=5, value=max_masa)
    val_masa.font = kpi_value_font
    val_masa.alignment = Alignment(horizontal="center", vertical="center")
    val_masa.fill = light_blue_fill
    val_masa.number_format = '#,##0'
    val_masa.border = thin_border
    ws.cell(row=5, column=6).border = thin_border

    ws.merge_cells("H4:I4")
    ws.cell(row=4, column=8, value="Fuerza Máx (N)").font = kpi_label_font
    ws.cell(row=4, column=8).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=8).fill = zebra_fill
    ws.cell(row=4, column=8).border = thin_border
    ws.cell(row=4, column=9).border = thin_border
    
    ws.merge_cells("H5:I5")
    val_fuerza = ws.cell(row=5, column=8, value=max_fuerza)
    val_fuerza.font = kpi_value_font
    val_fuerza.alignment = Alignment(horizontal="center", vertical="center")
    val_fuerza.fill = light_blue_fill
    val_fuerza.number_format = '#,##0'
    val_fuerza.border = thin_border
    ws.cell(row=5, column=9).border = thin_border

    # 3. Título de la Tabla (Fila 7)
    ws.cell(row=7, column=1, value="Historial Detallado de Ensayos").font = bold_dark_font
    ws.row_dimensions[7].height = 20

    # 4. Encabezados de la Tabla (Fila 8)
    headers = [
        "ID",
        "Fecha / Hora",
        "Escenario del Entorno",
        "Masa (kg)",
        "Fuerza Aplicada (N)",
        "Rampa / Clima",
        "Acel. Prom (m/s²)",
        "Tiempo Total (s)",
        "Desplazamiento (m)"
    ]
    
    ws.row_dimensions[8].height = 28
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = medium_blue_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 5. Escribir Filas de Datos (Fila 9 en adelante)
    for idx, r in enumerate(records, start=9):
        rampa_clima = r.clima or 'Seco'
        if r.escenario.lower() in ["automovil", "automotriz", "camion", "motocicleta"]:
            if r.inclinacion and abs(r.inclinacion) > 0.01:
                rampa_clima = f"{rampa_clima} ({r.inclinacion:+.1f}°)"
        elif r.escenario.lower() == "elevador":
            rampa_clima = r.clima or 'Aire'
        elif r.escenario.lower() == "avion":
            rampa_clima = r.clima or 'Pista Seca'
        elif r.escenario.lower() == "atwood":
            rampa_clima = r.perfil_fuerza or 'Constante'

        row_vals = [
            f"#{r.id}",
            r.fecha.strftime("%Y-%m-%d %H:%M:%S") if hasattr(r.fecha, "strftime") else str(r.fecha),
            r.escenario,
            r.masa,
            r.fuerza,
            rampa_clima,
            r.aceleracion_promedio,
            r.tiempo_total,
            r.distancia_recorrida
        ]
        
        ws.row_dimensions[idx].height = 20
        use_zebra = (idx % 2 == 0)
        
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill
                
            # Casing/Alineaciones y Formatos
            if col_idx == 1:
                cell.font = data_bold_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                cell.font = data_normal_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:
                cell.font = data_normal_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx in [4, 5]:
                cell.font = data_normal_font
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            elif col_idx == 6:
                cell.font = data_normal_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 7:
                cell.font = data_normal_font
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.000'
            elif col_idx in [8, 9]:
                cell.font = data_normal_font
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.00'

    # Ajustar anchos automáticamente
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.coordinate in ws.merged_cells:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # Manual overrides para un mejor espaciado visual
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['F'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


